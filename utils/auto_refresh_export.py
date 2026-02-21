"""
auto_refresh_export.py — IFControll v3.0
• Auto-refresh global de 15 segundos para todas as abas
• Correção de fuso horário (UTC-3 / Brasília)
• Exportação multi-formato: CSV, XLSX, XLS, PDF, TXT
• Ctrl+C universal em qualquer widget de texto/tabela

INTEGRAÇÃO:
  1. Importe no topo do arquivo principal.
  2. Substitua as funções now_str(), ts() e parse_dt() pelas daqui.
  3. Use auto_refresh_register() para registrar cada aba.
  4. Substitua mk_export_btn() pela versão abaixo.
"""

import csv, os, io, threading
from datetime import datetime, timedelta, timezone
from tkinter import filedialog, messagebox
import tkinter as tk

# ─── FUSO HORÁRIO BRASIL (UTC-3) ─────────────────────────────────────────────
TZ_BR = timezone(timedelta(hours=-3))

def now_br():
    """Retorna datetime atual no fuso UTC-3."""
    return datetime.now(TZ_BR)

def now_str():
    """Hora atual formatada no fuso brasileiro."""
    return now_br().strftime("%H:%M:%S")

def ts(dt):
    """Converte datetime para Unix timestamp.
    Se dt não tiver tzinfo, assume UTC-3."""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=TZ_BR)
    return int(dt.timestamp())

def parse_dt(s):
    """Parse de string de data/hora — igual ao original, sem alteração de tz."""
    for fmt in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(str(s).strip(), fmt)
        except:
            pass
    return None

def fmt_now_default():
    """Retorna string de 'agora' para preencher campos de data."""
    return now_br().strftime("%d/%m/%Y %H:%M")

def fmt_hours_ago(h=8):
    return (now_br() - timedelta(hours=h)).strftime("%d/%m/%Y %H:%M")

def fmt_days_ago(d=7):
    return (now_br() - timedelta(days=d)).strftime("%d/%m/%Y %H:%M")


# ─── AUTO-REFRESH GLOBAL ─────────────────────────────────────────────────────
_refresh_tasks = []   # lista de (name, callable)
_auto_enabled  = True
_INTERVAL_MS   = 60_000  # 60 segundos

def auto_refresh_register(name, fn):
    """Registra uma função de refresh para ser chamada automaticamente."""
    _refresh_tasks.append((name, fn))

def auto_refresh_run_all():
    """Executa todos os refreshes registrados em threads separadas."""
    for name, fn in _refresh_tasks:
        threading.Thread(target=fn, daemon=True, name=f"refresh-{name}").start()

def auto_refresh_loop(root):
    """Loop principal — chame uma vez após criar a janela."""
    if _auto_enabled:
        auto_refresh_run_all()
    root.after(_INTERVAL_MS, lambda: auto_refresh_loop(root))

def auto_refresh_set_enabled(val: bool):
    global _auto_enabled
    _auto_enabled = val


# ─── EXPORTAÇÃO MULTI-FORMATO ─────────────────────────────────────────────────
def _cols_rows_from_tree(tree):
    cols = [tree.heading(c)["text"] for c in tree["columns"]]
    rows = [tree.item(r)["values"] for r in tree.get_children()]
    return cols, rows

def _cols_rows_from_ftree(ft):
    """FilterableTree wrapper."""
    return _cols_rows_from_tree(ft.tree)

def export_universal(source, title="Exportar", source_type="tree"):
    """
    Exporta dados de um Treeview, FilterableTree ou Text widget
    nos formatos: CSV, XLSX, XLS, PDF, TXT.

    source_type: "tree" | "ftree" | "text"
    """
    # ── Coleta dados ──────────────────────────────────────────────
    if source_type in ("tree", "ftree"):
        if source_type == "ftree":
            cols, rows = _cols_rows_from_ftree(source)
        else:
            cols, rows = _cols_rows_from_tree(source)
        if not rows:
            messagebox.showinfo("Exportar", "Nenhum dado para exportar.")
            return
        text_content = None
    else:  # "text"
        text_content = source.get("1.0", "end").strip()
        cols, rows = [], []
        if not text_content:
            messagebox.showinfo("Exportar", "Nenhum conteúdo para exportar.")
            return

    # ── Diálogo de arquivo ────────────────────────────────────────
    ts_str = now_br().strftime("%Y%m%d_%H%M%S")
    default_name = f"ifcontroll_{ts_str}"

    filetypes = [
        ("CSV (separado por ;)", "*.csv"),
        ("Excel XLSX", "*.xlsx"),
        ("Excel XLS", "*.xls"),
        ("PDF", "*.pdf"),
        ("Texto TXT", "*.txt"),
        ("Todos os arquivos", "*.*"),
    ]

    path = filedialog.asksaveasfilename(
        title=title,
        filetypes=filetypes,
        defaultextension=".csv",
        initialfile=default_name,
    )
    if not path:
        return

    ext = os.path.splitext(path)[1].lower()

    try:
        if ext == ".csv":
            _export_csv(path, cols, rows, text_content)
        elif ext == ".xlsx":
            _export_xlsx(path, cols, rows, text_content)
        elif ext == ".xls":
            _export_xls(path, cols, rows, text_content)
        elif ext == ".pdf":
            _export_pdf(path, cols, rows, text_content, title)
        else:  # .txt ou qualquer outro
            _export_txt(path, cols, rows, text_content)

        messagebox.showinfo("Exportar", f"✔ Arquivo salvo:\n{path}")
    except ImportError as e:
        messagebox.showerror("Dependência faltando",
            f"Instale a biblioteca necessária:\n{e}\n\n"
            f"Execute: pip install openpyxl xlwt reportlab")
    except Exception as e:
        messagebox.showerror("Erro ao exportar", str(e))


# ── Formatos internos ──────────────────────────────────────────────
def _export_csv(path, cols, rows, text_content):
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        if text_content:
            f.write(text_content)
        else:
            w = csv.writer(f, delimiter=";")
            w.writerow(cols)
            w.writerows(rows)

def _export_txt(path, cols, rows, text_content):
    with open(path, "w", encoding="utf-8") as f:
        if text_content:
            f.write(text_content)
        else:
            widths = [max(len(str(c)), max((len(str(r[i])) for r in rows if i < len(r)), default=0))
                      for i, c in enumerate(cols)]
            header = "  ".join(str(c).ljust(widths[i]) for i, c in enumerate(cols))
            f.write(header + "\n")
            f.write("─" * len(header) + "\n")
            for row in rows:
                f.write("  ".join(str(v).ljust(widths[i]) for i, v in enumerate(row)) + "\n")

def _export_xlsx(path, cols, rows, text_content):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl não instalado. Execute: pip install openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IFControll"

    if text_content:
        for i, line in enumerate(text_content.split("\n"), 1):
            ws.cell(row=i, column=1, value=line)
    else:
        # Cabeçalho
        hdr_fill = PatternFill("solid", fgColor="1E2335")
        hdr_font = Font(bold=True, color="00C8F8")
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = hdr_fill
            cell.font = hdr_font
            cell.alignment = Alignment(horizontal="left")

        # Dados
        alt_fill = PatternFill("solid", fgColor="181C29")
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate(row, 1):
                cell = ws.cell(row=ri, column=ci, value=str(val) if val is not None else "")
                if ri % 2 == 0:
                    cell.fill = alt_fill
                cell.alignment = Alignment(horizontal="left")

        # Auto-width
        for col_cells in ws.columns:
            length = max(len(str(c.value or "")) for c in col_cells)
            ws.column_dimensions[col_cells[0].column_letter].width = min(length + 4, 40)

    wb.save(path)

def _export_xls(path, cols, rows, text_content):
    try:
        import xlwt
    except ImportError:
        raise ImportError("xlwt não instalado. Execute: pip install xlwt")

    wb = xlwt.Workbook(encoding="utf-8")
    ws = wb.add_sheet("IFControll")

    hdr_style = xlwt.easyxf(
        "font: bold true, colour white; "
        "pattern: pattern solid, fore_colour dark_blue_ega; "
        "align: horiz left"
    )
    data_style = xlwt.easyxf("align: horiz left")

    if text_content:
        for ri, line in enumerate(text_content.split("\n")):
            ws.write(ri, 0, line, data_style)
    else:
        for ci, col in enumerate(cols):
            ws.write(0, ci, col, hdr_style)
        for ri, row in enumerate(rows, 1):
            for ci, val in enumerate(row):
                ws.write(ri, ci, str(val) if val is not None else "", data_style)

    wb.save(path)

def _export_pdf(path, cols, rows, text_content, title="Relatório"):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Preformatted
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import cm
    except ImportError:
        raise ImportError("reportlab não instalado. Execute: pip install reportlab")

    is_landscape = cols and len(cols) > 6
    pagesize = landscape(A4) if is_landscape else A4
    doc = SimpleDocTemplate(path, pagesize=pagesize,
                            leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"],
                                 textColor=colors.HexColor("#00C8F8"),
                                 fontSize=14, spaceAfter=6)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"],
                               textColor=colors.HexColor("#8B93B5"),
                               fontSize=8, spaceAfter=12)

    story = []
    story.append(Paragraph(title or "IFControll — Relatório", title_style))
    story.append(Paragraph(
        f"Gerado em: {now_br().strftime('%d/%m/%Y %H:%M:%S')} (Brasília UTC-3)",
        sub_style))
    story.append(Spacer(1, 0.3*cm))

    if text_content:
        mono = ParagraphStyle("mono", parent=styles["Code"],
                              fontSize=7, leading=10,
                              textColor=colors.HexColor("#DDE1F0"),
                              backColor=colors.HexColor("#12151E"))
        story.append(Preformatted(text_content, mono))
    else:
        # Tabela
        data = [cols] + [[str(v) if v is not None else "" for v in row] for row in rows]
        col_count = len(cols)
        pw = (landscape(A4)[0] if is_landscape else A4[0]) - 2*cm

        # Largura das colunas distribuída
        col_w = [pw / col_count] * col_count

        tbl = Table(data, colWidths=col_w, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1E2335")),
            ("TEXTCOLOR",  (0,0), (-1,0), colors.HexColor("#00C8F8")),
            ("FONTNAME",   (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE",   (0,0), (-1,0), 8),
            ("FONTSIZE",   (0,1), (-1,-1), 7),
            ("BACKGROUND", (0,1), (-1,-1), colors.HexColor("#181C29")),
            ("TEXTCOLOR",  (0,1), (-1,-1), colors.HexColor("#DDE1F0")),
            ("ROWBACKGROUNDS", (0,1), (-1,-1),
             [colors.HexColor("#181C29"), colors.HexColor("#12151E")]),
            ("GRID", (0,0), (-1,-1), 0.25, colors.HexColor("#232840")),
            ("ALIGN", (0,0), (-1,-1), "LEFT"),
            ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
            ("TOPPADDING", (0,0), (-1,-1), 3),
            ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ]))
        story.append(tbl)

    story.append(Spacer(1, 0.5*cm))
    footer_style = ParagraphStyle("foot", parent=styles["Normal"],
                                  fontSize=7, textColor=colors.HexColor("#58607A"))
    story.append(Paragraph("IFControll v3.0 · Powered by Fulltrack2 REST API", footer_style))

    doc.build(story)


# ─── BOTÃO DE EXPORTAÇÃO UNIVERSAL ────────────────────────────────────────────
def mk_export_btn(parent, source, source_type="tree", title="Exportar"):
    """
    Substitui o mk_export_btn original.
    source_type: "tree" | "ftree" | "text"
    """
    import tkinter as tk
    # Import lazy do C para pegar o tema atual
    def _get_C():
        try:
            from theme_manager import C
            return C
        except:
            return {"surface3":"#1E2335","accent":"#00C8F8","bg":"#0B0D12"}

    def do_export():
        C = _get_C()
        export_universal(source, title=title, source_type=source_type)

    b = tk.Label(parent, text="📥 EXPORTAR",
                 bg="#1E2335", fg="#00C8F8",
                 font=("Helvetica Neue", 9, "bold"),
                 padx=10, pady=5, cursor="hand2", relief="flat")
    b.bind("<Button-1>", lambda e: do_export())
    return b


# ─── CTRL+C UNIVERSAL ─────────────────────────────────────────────────────────
def bind_global_copy(root):
    """
    Vincula Ctrl+C global na janela raiz.
    Funciona para Entry, Text, Treeview e Labels selecionáveis.
    """
    def _copy(event):
        widget = root.focus_get()
        if widget is None:
            return
        wtype = widget.winfo_class()

        try:
            if wtype == "Entry":
                if widget.selection_present():
                    text = widget.selection_get()
                else:
                    text = widget.get()
                root.clipboard_clear()
                root.clipboard_append(text)

            elif wtype == "Text":
                try:
                    text = widget.get(tk.SEL_FIRST, tk.SEL_LAST)
                except tk.TclError:
                    text = widget.get("1.0", "end").strip()
                root.clipboard_clear()
                root.clipboard_append(text)

            elif wtype == "Treeview":
                sel = widget.selection()
                if sel:
                    lines = ["\t".join(str(v) for v in widget.item(i)["values"])
                             for i in sel]
                    root.clipboard_clear()
                    root.clipboard_append("\n".join(lines))
        except Exception:
            pass

    root.bind_all("<Control-c>", _copy, add="+")
    root.bind_all("<Control-C>", _copy, add="+")


# ─── WIDGET DE CONTROLE DE AUTO-REFRESH ───────────────────────────────────────
def mk_refresh_controls(parent, root):
    """
    Cria um mini-painel com: checkbox auto-refresh, label de intervalo,
    e botão 'Atualizar Tudo Agora'.
    Cole no header ou footer da janela principal.
    """
    f = tk.Frame(parent, bg="#12151E")

    auto_var = tk.BooleanVar(value=True)

    def toggle_auto():
        auto_refresh_set_enabled(auto_var.get())

    chk = tk.Checkbutton(f, text="Auto 60s", variable=auto_var,
                         command=toggle_auto,
                         bg="#12151E", fg="#8B93B5",
                         activebackground="#12151E",
                         selectcolor="#1E2335",
                         font=("Helvetica Neue", 8))
    chk.pack(side="left", padx=(0, 6))

    status_lbl = tk.Label(f, text="", bg="#12151E", fg="#58607A",
                          font=("Helvetica Neue", 7))
    status_lbl.pack(side="left")

    def _update_now():
        status_lbl.config(text="⟳ atualizando...")
        auto_refresh_run_all()
        status_lbl.config(text=f"✔ {now_str()}")

    btn_now = tk.Label(f, text="⟳ Tudo agora",
                       bg="#1E2335", fg="#00C8F8",
                       font=("Helvetica Neue", 8, "bold"),
                       padx=8, pady=3, cursor="hand2")
    btn_now.bind("<Button-1>", lambda e: threading.Thread(target=_update_now, daemon=True).start())
    btn_now.pack(side="left", padx=4)

    return f
